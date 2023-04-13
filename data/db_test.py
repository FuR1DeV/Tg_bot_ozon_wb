import asyncio
from datetime import datetime

from sqlalchemy import and_, or_

from bot import dp
from data.commands import general_set
from settings import config
from data.db_gino import db
from data import db_gino


async def db_test():
    await db_gino.on_startup(dp)
    from openpyxl import load_workbook

    book = load_workbook(filename=f"Бот (2).xlsx")
    sheet = book["Лист1"]

    for row in range(2, sheet.max_row + 1):
        if sheet["A" + str(row)].value is not None:
            await general_set.product_ozon_add(sheet["A" + str(row)].value,
                                               sheet["B" + str(row)].value,
                                               sheet["C" + str(row)].value,
                                               sheet["D" + str(row)].value,
                                               sheet["E" + str(row)].value,
                                               sheet["F" + str(row)].value)


if __name__ == '__main__':
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(db_test())